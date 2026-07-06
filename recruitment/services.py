"""Business logic for the recruitment app: duplicate detection, status
transitions, Excel export, and the shared Excel-import routine used by both
`manage.py import_candidates` and `recruitment/sharepoint_sync.py`.
"""
import io

import openpyxl
from django.utils import timezone

from .models import Candidate, CandidateStatusHistory, Role

STATUS = Candidate.Status

# Tabs shown on the dashboard, in display order.
DASHBOARD_TABS = [
    ('central', 'Central Repository', None),
    ('under_review', 'Under Review', STATUS.UNDER_REVIEW),
    ('shortlisted', 'Shortlisted', STATUS.SHORTLISTED),
    ('rejected', 'Rejected', STATUS.REJECTED),
    ('blacklisted', 'Blacklisted', STATUS.BLACKLISTED),
]


def determine_initial_status(email):
    """Duplicate-detection rule applied whenever a brand-new candidate row
    is created (manually or via import).

    - email already exists and that candidate is BLACKLISTED -> BLACKLISTED
    - email already exists (any other status)                -> REAPPLY
    - otherwise                                               -> NEW
    """
    existing = Candidate.objects.filter(email__iexact=email)
    if not existing.exists():
        return STATUS.NEW, False
    if existing.filter(current_status=STATUS.BLACKLISTED).exists():
        return STATUS.BLACKLISTED, True
    return STATUS.REAPPLY, True


def change_status(candidate, new_status, user=None, blacklist_reason=None):
    """Move a candidate to a new status and record it in the audit trail."""
    old_status = candidate.current_status
    if old_status == new_status:
        return candidate
    candidate.current_status = new_status
    if new_status == STATUS.BLACKLISTED and blacklist_reason:
        candidate.blacklist_reason = blacklist_reason
    candidate.save(update_fields=['current_status', 'blacklist_reason', 'updated_at'])
    CandidateStatusHistory.objects.create(
        candidate=candidate,
        old_status=old_status,
        new_status=new_status,
        changed_by=user,
    )
    return candidate


def record_creation(candidate, user=None):
    """Log the initial status assignment in the audit trail."""
    CandidateStatusHistory.objects.create(
        candidate=candidate,
        old_status='',
        new_status=candidate.current_status,
        changed_by=user,
    )


def bulk_change_status(candidate_ids, new_status, user=None):
    candidates = Candidate.objects.filter(pk__in=candidate_ids)
    count = 0
    for candidate in candidates:
        change_status(candidate, new_status, user=user)
        count += 1
    return count


def export_candidates_xlsx(queryset):
    """Build an in-memory .xlsx workbook for the given Candidate queryset."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Candidates'
    headers = [
        'Name', 'Email', 'Phone', 'Role', 'Applied Date', 'Status',
        'Duplicate', 'Experience', 'Notice Period', 'Expected Salary',
        'Source', 'Remarks',
    ]
    ws.append(headers)
    for c in queryset:
        ws.append([
            c.full_name,
            c.email,
            c.phone or '',
            c.role.title if c.role else '',
            timezone.localtime(c.applied_at).strftime('%Y-%m-%d %H:%M') if c.applied_at else '',
            c.get_current_status_display(),
            'Yes' if c.duplicate_flag else 'No',
            str(c.experience) if c.experience is not None else '',
            c.notice_period or '',
            c.expected_salary or '',
            c.source or '',
            c.remarks or '',
        ])
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Excel import (SharePoint export) — shared by the management command and
# the SharePoint sync utility.
# ---------------------------------------------------------------------------

ACTIVE_TRUE_VALUES = {'yes', 'open', 'active', 'true', '1'}

# The "Central Sheet" columns (0-indexed) as exported from SharePoint.
CENTRAL_SHEET_COLUMNS = [
    'sl_no', 'mail_date', 'name', 'mail_id', 'mobile_number', 'role_applied',
    'education', 'cv_link', 'hyper_link', 'role_selected', 'status',
    'is_copied', 'key', 'source', 'remarks', 'id', 'screened_on', 'screened_by',
]

ROLE_SHEET_COLUMNS = ['role', 'shortform', 'opening_date', 'closing_date', 'active', 'id', 'requirement']


def _row_dict(row, columns):
    return dict(zip(columns, row))


def _map_status(status_text, is_copied_text):
    """Translate the legacy 'Status' / 'Is Copied' sheet columns into a
    single Candidate.Status value, per the "ignore sheet separation, use
    status mapping" instruction.
    """
    is_copied = (is_copied_text or '').strip().lower()
    status = (status_text or '').strip().lower()

    if is_copied == 'copiedtoshortlist' or status == 'shortlisted':
        return STATUS.SHORTLISTED
    if is_copied == 'copiedtoreview':
        return STATUS.UNDER_REVIEW
    if status in ('rejected', 'not shortlisted', 'not applicable'):
        return STATUS.REJECTED
    return None  # unresolved -> fall back to duplicate-detection rule


def import_roles(ws):
    """Import the ROLE sheet. Returns (created, updated)."""
    created = updated = 0
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        data = _row_dict(row, ROLE_SHEET_COLUMNS)
        title = (data.get('role') or '').strip()
        if not title:
            continue
        is_active = str(data.get('active') or '').strip().lower() in ACTIVE_TRUE_VALUES
        role, made = Role.objects.get_or_create(
            title__iexact=title,
            defaults={'title': title, 'is_active': is_active},
        )
        if made:
            created += 1
        else:
            if role.is_active != is_active:
                role.is_active = is_active
                role.save(update_fields=['is_active'])
                updated += 1
    return created, updated


def _resolve_role(role_name):
    role_name = (role_name or '').strip()
    if not role_name:
        return None
    role = Role.objects.filter(title__iexact=role_name).first()
    if role:
        return role
    return Role.objects.create(title=role_name, is_active=True)


def import_candidates(ws, user=None):
    """Import the Central Sheet. Returns dict with created/skipped counts."""
    created = 0
    skipped = 0
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        data = _row_dict(row, CENTRAL_SHEET_COLUMNS)
        name = (data.get('name') or '').strip()
        email = (data.get('mail_id') or '').strip()
        if not name or not email:
            skipped += 1
            continue

        source = (data.get('source') or '').strip() or None

        # Idempotency: re-running the import should not duplicate rows
        # already pulled in from a previous run.
        if Candidate.objects.filter(email__iexact=email, full_name__iexact=name, source=source).exists():
            skipped += 1
            continue

        role = _resolve_role(data.get('role_selected') or data.get('role_applied'))
        mapped_status = _map_status(data.get('status'), data.get('is_copied'))
        fallback_status, duplicate_flag = determine_initial_status(email)
        status = mapped_status if mapped_status is not None else fallback_status

        candidate = Candidate.objects.create(
            full_name=name,
            email=email,
            phone=(data.get('mobile_number') or '').strip() or None,
            role=role,
            current_status=status,
            source=source,
            remarks=(data.get('remarks') or '').strip() or None,
            duplicate_flag=duplicate_flag,
        )
        mail_date = data.get('mail_date')
        if mail_date:
            if timezone.is_naive(mail_date):
                mail_date = timezone.make_aware(mail_date)
            Candidate.objects.filter(pk=candidate.pk).update(applied_at=mail_date)
        record_creation(candidate, user=user)
        created += 1
    return {'created': created, 'skipped': skipped}


def import_workbook(filepath, user=None):
    """Open a Central-Repository-style workbook and import Roles + Candidates."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {'roles_created': 0, 'roles_updated': 0, 'candidates_created': 0, 'candidates_skipped': 0}

    if 'ROLE' in wb.sheetnames:
        result['roles_created'], result['roles_updated'] = import_roles(wb['ROLE'])

    sheet_name = 'Central Sheet' if 'Central Sheet' in wb.sheetnames else wb.sheetnames[0]
    candidate_result = import_candidates(wb[sheet_name], user=user)
    result['candidates_created'] = candidate_result['created']
    result['candidates_skipped'] = candidate_result['skipped']
    return result
